import logging
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional, List
from app.database import get_db
from app.routers.admin import require_admin
from datetime import date, datetime
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/admin/facturacion",
    tags=["facturacion"]
)


class PagoManual(BaseModel):
    usuario_id: int
    plan_id: int
    monto: float
    metodo: str
    descripcion: Optional[str] = None


# ─── Pagos ────────────────────────────────────────────────────────────────────

@router.get("/pagos")
def listar_pagos(
    estado: Optional[str] = Query(None),
    pasarela: Optional[str] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: int = Depends(require_admin)
):
    try:
        condiciones = ["1=1"]
        params: dict = {}

        if estado:
            condiciones.append("p.estado = :estado")
            params["estado"] = estado
        if pasarela:
            condiciones.append("p.pasarela = :pasarela")
            params["pasarela"] = pasarela
        if fecha_desde:
            condiciones.append("p.created_at >= :fecha_desde")
            params["fecha_desde"] = fecha_desde
        if fecha_hasta:
            condiciones.append("p.created_at <= :fecha_hasta")
            params["fecha_hasta"] = fecha_hasta

        where = " AND ".join(condiciones)

        rows = db.execute(text(f"""
            SELECT p.id,
                   u.nombre || ' ' || u.apellido AS usuario_nombre,
                   u.email AS usuario_email,
                   p.monto, p.moneda, p.pasarela, p.estado, p.tipo,
                   p.fecha_aprobacion, p.created_at
            FROM pagos p
            JOIN usuarios u ON u.id = p.usuario_id
            WHERE {where}
            ORDER BY p.created_at DESC
        """), params).fetchall()

        return [
            {
                "id": r.id,
                "usuario_nombre": r.usuario_nombre,
                "usuario_email": r.usuario_email,
                "monto": float(r.monto) if r.monto is not None else None,
                "moneda": r.moneda,
                "pasarela": r.pasarela,
                "estado": r.estado,
                "tipo": r.tipo,
                "fecha": r.fecha_aprobacion.isoformat() if r.fecha_aprobacion else r.created_at.isoformat(),
                "fecha_aprobacion": r.fecha_aprobacion.isoformat() if r.fecha_aprobacion else None,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error interno: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


# ─── Resumen facturación ───────────────────────────────────────────────────────

@router.get("/resumen")
def resumen_facturacion(
    db: Session = Depends(get_db),
    _: int = Depends(require_admin)
):
    try:
        mes_actual = db.execute(text("""
            SELECT
                COALESCE(SUM(monto) FILTER (WHERE estado = 'aprobado'), 0) AS total,
                COUNT(*) FILTER (WHERE estado = 'aprobado') AS aprobados,
                COUNT(*) FILTER (WHERE estado = 'rechazado') AS rechazados
            FROM pagos
            WHERE created_at >= DATE_TRUNC('month', NOW())
        """)).fetchone()

        mes_anterior = db.execute(text("""
            SELECT COALESCE(SUM(monto), 0) AS total FROM pagos
            WHERE estado = 'aprobado'
            AND created_at >= DATE_TRUNC('month', NOW() - INTERVAL '1 month')
            AND created_at < DATE_TRUNC('month', NOW())
        """)).fetchone()

        total_actual = float(mes_actual.total)
        total_ant = float(mes_anterior.total)
        variacion = round((total_actual - total_ant) / total_ant * 100, 2) if total_ant > 0 else 0

        por_pasarela = db.execute(text("""
            SELECT pasarela,
                   COALESCE(SUM(monto), 0) AS total,
                   COUNT(*) AS cantidad
            FROM pagos
            WHERE estado = 'aprobado'
            AND created_at >= DATE_TRUNC('month', NOW())
            GROUP BY pasarela
            ORDER BY total DESC
        """)).fetchall()

        return {
            "total_mes": total_actual,
            "total_mes_anterior": total_ant,
            "variacion_porcentual": variacion,
            "pagos_aprobados": mes_actual.aprobados,
            "pagos_rechazados": mes_actual.rechazados,
            "por_pasarela": [
                {"pasarela": r.pasarela, "total": float(r.total), "cantidad": r.cantidad}
                for r in por_pasarela
            ],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error interno: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


# ─── Facturas ─────────────────────────────────────────────────────────────────

@router.get("/facturas")
def listar_facturas(
    tipo: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: int = Depends(require_admin)
):
    try:
        condiciones = ["1=1"]
        params: dict = {}

        if tipo:
            condiciones.append("f.tipo = :tipo")
            params["tipo"] = tipo
        if mes:
            condiciones.append("TO_CHAR(f.fecha_emision, 'YYYY-MM') = :mes")
            params["mes"] = mes

        where = " AND ".join(condiciones)

        rows = db.execute(text(f"""
            SELECT id, numero_factura, tipo, razon_social,
                   monto_total, estado, fecha_emision
            FROM facturas f
            WHERE {where}
            ORDER BY fecha_emision DESC
        """), params).fetchall()

        return [
            {
                "id": r.id,
                "numero_factura": r.numero_factura,
                "tipo": r.tipo,
                "razon_social": r.razon_social,
                "monto_total": float(r.monto_total) if r.monto_total is not None else None,
                "estado": r.estado,
                "fecha_emision": r.fecha_emision.isoformat() if r.fecha_emision else None,
            }
            for r in rows
        ]
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error interno: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


# ─── Pago manual ──────────────────────────────────────────────────────────────

@router.post("/pagos/manual")
def pago_manual(
    datos: PagoManual,
    db: Session = Depends(get_db),
    _: int = Depends(require_admin)
):
    try:
        plan = db.execute(
            text("SELECT id, precio_mensual FROM planes WHERE id = :id AND activo = true"),
            {"id": datos.plan_id}
        ).fetchone()
        if not plan:
            raise HTTPException(status_code=404, detail="Plan no encontrado o inactivo")

        pago = db.execute(text("""
            INSERT INTO pagos (usuario_id, monto, pasarela, estado, tipo, fecha_aprobacion)
            VALUES (:usuario_id, :monto, 'manual', 'aprobado', :tipo, NOW())
            RETURNING id
        """), {
            "usuario_id": datos.usuario_id,
            "monto": datos.monto,
            "tipo": datos.descripcion or "pago_manual",
        }).fetchone()

        suscripcion_pendiente = db.execute(text("""
            SELECT id FROM suscripciones
            WHERE usuario_id = :uid AND estado = 'pendiente_pago'
            ORDER BY created_at DESC LIMIT 1
        """), {"uid": datos.usuario_id}).fetchone()

        suscripcion_id = None
        if suscripcion_pendiente:
            db.execute(
                text("UPDATE suscripciones SET estado = 'activa' WHERE id = :id"),
                {"id": suscripcion_pendiente.id}
            )
            suscripcion_id = suscripcion_pendiente.id
        else:
            nueva = db.execute(text("""
                INSERT INTO suscripciones
                  (usuario_id, plan_id, estado, fecha_inicio, precio_pagado)
                VALUES (:uid, :plan_id, 'activa', CURRENT_DATE, :precio)
                RETURNING id
            """), {
                "uid": datos.usuario_id,
                "plan_id": datos.plan_id,
                "precio": datos.monto,
            }).fetchone()
            suscripcion_id = nueva.id

        db.commit()
        return {
            "pago_id": pago.id,
            "suscripcion_id": suscripcion_id,
            "mensaje": "Pago manual registrado y suscripción activada correctamente",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error interno: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


# ─── Exportar Mediquo ──────────────────────────────────────────────────────────

@router.get("/exportar-mediquo")
def exportar_mediquo(
    db: Session = Depends(get_db),
    _: int = Depends(require_admin)
):
    try:
        rows = db.execute(text("""
            SELECT u.nombre, u.apellido, u.email, u.telefono, u.dni,
                   u.fecha_nacimiento, p.nombre AS plan_nombre,
                   s.estado, s.fecha_inicio
            FROM suscripciones s
            JOIN usuarios u ON u.id = s.usuario_id
            JOIN planes p ON p.id = s.plan_id
            WHERE DATE(s.created_at) = CURRENT_DATE
              AND s.estado = 'pendiente_pago'
            ORDER BY s.created_at DESC
        """)).fetchall()

        fecha_hoy = date.today().isoformat()
        ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suscriptores"

        # Fila 1: título
        ws.merge_cells("A1:I1")
        celda_titulo = ws["A1"]
        celda_titulo.value = f"Reporte CelDoctor para Mediquo — {fecha_hoy}"
        celda_titulo.font = Font(bold=True, color="FFFFFF", size=13)
        celda_titulo.fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

        # Fila 2: metadata
        ws.merge_cells("A2:I2")
        celda_meta = ws["A2"]
        celda_meta.value = f"Generado: {ahora} | Total: {len(rows)} registros"
        celda_meta.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        celda_meta.font = Font(italic=True)

        # Fila 3: headers
        headers = ["Nombre", "Apellido", "Email", "Teléfono", "DNI",
                   "Fecha Nacimiento", "Plan", "Estado", "Fecha Suscripción"]
        header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            celda = ws.cell(row=3, column=col_idx, value=header)
            celda.fill = header_fill
            celda.font = header_font

        # Fila 4+: datos
        if rows:
            for r in rows:
                ws.append([
                    r.nombre, r.apellido, r.email, r.telefono,
                    r.dni if r.dni is not None else "",
                    r.fecha_nacimiento.isoformat() if r.fecha_nacimiento else "",
                    r.plan_nombre, r.estado,
                    r.fecha_inicio.isoformat() if r.fecha_inicio else "",
                ])
        else:
            ws.merge_cells("A4:I4")
            ws["A4"].value = "Sin nuevos suscriptores hoy"

        # Autoajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"mediquo_{fecha_hoy}_{len(rows)}_registros.xlsx"
        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error interno: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


# ─── Marcar exportados ────────────────────────────────────────────────────────

class MarcarExportados(BaseModel):
    suscripcion_ids: List[int]


@router.post("/marcar-exportados")
def marcar_exportados(
    datos: MarcarExportados,
    db: Session = Depends(get_db),
    _: int = Depends(require_admin)
):
    try:
        db.execute(text("""
            INSERT INTO auditoria (accion, tabla_afectada, datos_nuevos)
            VALUES (:accion, :tabla, :datos)
        """), {
            "accion": "exportado_a_mediquo",
            "tabla": "suscripciones",
            "datos": json.dumps({
                "suscripcion_ids": datos.suscripcion_ids,
                "fecha": date.today().isoformat()
            }),
        })
        db.commit()
        return {"registrado": True, "cantidad": len(datos.suscripcion_ids)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error interno: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


# ─── Historial exportaciones ───────────────────────────────────────────────────

@router.get("/historial-exportaciones")
def historial_exportaciones(
    _: int = Depends(require_admin)
):
    return []
