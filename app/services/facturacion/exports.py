from datetime import date, datetime
import io
import json
import logging

from fastapi import HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.schemas.facturacion import MarcarExportados

logger = logging.getLogger(__name__)


def exportar_mediquo(db: Session):
    try:
        rows = db.execute(text("""
            SELECT s.id AS suscripcion_id,
                   u.nombre, u.apellido, u.email, u.telefono, u.dni,
                   u.fecha_nacimiento, p.nombre AS plan_nombre,
                   s.estado, s.fecha_inicio
            FROM suscripciones s
            JOIN usuarios u ON u.id = s.usuario_id
            JOIN planes p ON p.id = s.plan_id
            WHERE DATE(s.created_at) = CURRENT_DATE
              AND s.estado = 'pendiente_pago'
            ORDER BY s.created_at DESC
        """)).fetchall()

        fecha_hoy = date.today().isoformat()
        generado_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        workbook = _build_mediquo_workbook(rows, fecha_hoy, generado_en)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = f"mediquo_{fecha_hoy}_{len(rows)}_registros.xlsx"
        subscription_ids = ",".join(str(row.suscripcion_id) for row in rows)

        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Subscription-Ids": subscription_ids,
            },
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error interno: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


def marcar_exportados(
    db: Session,
    datos: MarcarExportados,
):
    try:
        if not datos.suscripcion_ids:
            raise HTTPException(status_code=400, detail="No se enviaron suscripciones para marcar")

        db.execute(text("""
            INSERT INTO auditoria (accion, tabla_afectada, datos_nuevos)
            VALUES (:accion, :tabla, :datos)
        """), {
            "accion": "exportado_a_mediquo",
            "tabla": "suscripciones",
            "datos": json.dumps({
                "suscripcion_ids": datos.suscripcion_ids,
                "fecha": date.today().isoformat(),
            }),
        })
        db.commit()

        return {
            "registrado": True,
            "cantidad": len(datos.suscripcion_ids),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error interno: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


def historial_exportaciones(db: Session):
    try:
        rows = db.execute(text("""
            SELECT datos_nuevos, created_at
            FROM auditoria
            WHERE tabla_afectada = 'suscripciones'
              AND accion = 'exportado_a_mediquo'
            ORDER BY created_at DESC
            LIMIT 50
        """)).fetchall()

        payload = []
        for row in rows:
            datos = row.datos_nuevos
            if isinstance(datos, str):
                datos = json.loads(datos)

            suscripcion_ids = (datos or {}).get("suscripcion_ids", [])
            payload.append({
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "suscripcion_ids": suscripcion_ids,
                "cantidad": len(suscripcion_ids),
                "fecha": (datos or {}).get("fecha"),
            })

        return payload
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error interno: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor.")


def _build_mediquo_workbook(rows, fecha_hoy: str, generado_en: str) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suscriptores"

    headers = [
        "Nombre",
        "Apellido",
        "Email",
        "Telefono",
        "DNI",
        "Fecha Nacimiento",
        "Plan",
        "Estado",
        "Fecha Suscripcion",
    ]

    sheet.merge_cells("A1:I1")
    title_cell = sheet["A1"]
    title_cell.value = f"Reporte CelDoctor para Mediquo - {fecha_hoy}"
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A2:I2")
    meta_cell = sheet["A2"]
    meta_cell.value = f"Generado: {generado_en} | Total: {len(rows)} registros"
    meta_cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    meta_cell.font = Font(italic=True)

    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    header_font = Font(bold=True)
    for index, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font

    if rows:
        for row in rows:
            sheet.append([
                row.nombre,
                row.apellido,
                row.email,
                row.telefono,
                row.dni if row.dni is not None else "",
                row.fecha_nacimiento.isoformat() if row.fecha_nacimiento else "",
                row.plan_nombre,
                row.estado,
                row.fecha_inicio.isoformat() if row.fecha_inicio else "",
            ])
    else:
        sheet.merge_cells("A4:I4")
        sheet["A4"].value = "Sin nuevos suscriptores hoy"

    for column_index in range(1, len(headers) + 1):
        max_length = len(headers[column_index - 1])
        column_letter = get_column_letter(column_index)
        for row_index in range(4, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value:
                max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 40)

    return workbook
